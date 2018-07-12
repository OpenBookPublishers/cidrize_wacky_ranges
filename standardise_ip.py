#!/usr/bin/env python2

# standardise-ip, by Chuan Tan <ct538@cam.ac.uk>
#
# Copyright (C) Chuan Tan 2018
#
# This programme is free software; you may redistribute and/or modify
# it under the terms of the Apache Licence, version 2.0.

# Reads from excel spreadsheet and converts them into IPNetwork, then outputs them
# usage: standardise_ip [inputfile] [sheetname] [columnnumber]
# You can easily check if an IPAddress is in IPNetwork by using
# If IPAddress in IPNetwork:
#           do whatever


from cidrize import cidrize,CidrizeError
from netaddr import *
from openpyxl import load_workbook
import re
import sys

# Global variables
BAD = 0
GOOD = 1

# Patterns

# 192.168.*.*
pat_privateip = re.compile("192\.168\..*\..*");

# match xx.xxx.xxx.xx-xx.xxx.xx.xx
pat_hyphen = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}");

# (xx.xx.xx.xx/xx)
pat_bracket = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}");

# xx.xxx.xx.*
pat_wildcard = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\*");

# xx.xxx.xx-xx.*
pat_thirdoctet_wildcard = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}\.\*");

# xx.xxx.xx-xx.xx
pat_thirdoctet = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}\.\d{1,3}");

# EZProxyIP123.405.2.3
pat_words = re.compile("[^0-9.-/*]");

# xx.xxx.(xx-xx).(xx-xx)
pat_two_brackets = re.compile("\d{1,3}\.\d{1,3}\.\(\d{1,3}-\d{1,3}\)\.\(\d{1,3}-\d{1,3}\)")

# xx.xxx.xx.xx
pat_simple = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}")

# xx.xxx.xx.xx-xxx
pat_fourthoctet = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}")

# xx.xxx.xxx.x[0-255]
pat_squarebrackets = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1}\[\d{1,3}-\d{1,3}\]");

# xx.xxx.xxx.[xxxx]
pat_squarebrackets_fourth = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\[\d{1,4}]");




# Dictionary of stuff to replace. For standardising.
def replace_all(text, dic):
    assert type(text) == type(u"")

    for i, j in dic.iteritems():
        text = text.replace(i, j)
    return text

reps = {
    "to": "-",
    "and": "&",
    u"\u002D": "-",
    u"\u058A": "-",
    u"\u05BE": "-",
    u"\u1400": "-",
    u"\u1806": "-",
    u"\u2011": "-",
    u"\u2012": "-",
    u"\u2013": "-",
    u"\u2014": "-",
    u"\u2015": "-",
    u"\u2E3A": "-",
    u"\u2E3B": "-",
    u"\uFE58": "-",
    u"\uFE63": "-",
    u"\uFF0D": "-",
}

def third_octet(ip_result):
    prefix = re.search(re.compile("\d{1,3}\.\d{1,3}"), ip_result).group(0)
    suffix = re.split(re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}-\d{1,3}"),
                      ip_result)[1]
    ip_range = re.search(re.compile("\d{1,3}-\d{1,3}"), ip_result).group(0)
    first = ip_range.split("-")[0]
    second = ip_range.split("-")[1]

    if suffix == ".*":
        startip = prefix + "." + first  + ".0"
        endip = prefix + "." + second + ".255"
    else:
        startip = prefix + "." + first  + suffix
        endip = prefix + "." + second  + suffix

    return iprange_to_cidrs(startip, endip)

def two_brackets(ip_result):
    prefix = re.search(re.compile("\d{1,3}\.\d{1,3}"), ip_result).group(0)
    third_octet = re.findall("\d{1,3}-\d{1,3}", ip_result)[0]
    fourth_octet = re.findall("\d{1,3}-\d{1,3}", ip_result)[1]

    third_octet_first = third_octet.split("-")[0]
    third_octet_second = third_octet.split("-")[1]

    fourth_octet_first = fourth_octet.split("-")[0]
    fourth_octet_second = fourth_octet.split("-")[1]

    startip = prefix + "." + third_octet_first + "." + fourth_octet_first
    endip = prefix + "." + third_octet_second + "." + fourth_octet_second

    return iprange_to_cidrs(startip, endip)

def handle_hyphenated_range(ip_result):
    startip, endip = ip_result.split("-")
    return iprange_to_cidrs(startip, endip)

def empty_list(anything):
    return []

class CategorisationError(Exception):
    pass

def categorise_individual_ip(ip):
    '''
    Take an individual IP address range expression `ip' and attempt to
    extract the net ranges from it.

    Return type: a list of netaddr.ip.IPNetwork's

    Example input: "41.220.19.209-41.220.19.222"

    which should have return value:

      [IPNetwork('41.220.19.209/32'),
       IPNetwork('41.220.19.210/31'),
       IPNetwork('41.220.19.212/30'),
       IPNetwork('41.220.19.216/30'),
       IPNetwork('41.220.19.220/31'),
       IPNetwork('41.220.19.222/32')]

    If `ip' doesn't match any known pattern, then throw CategorisationError.
    '''
    ip = ip.replace(" ","")

    pattern_handlers = [
        (pat_privateip, empty_list),
        (pat_hyphen, handle_hyphenated_range),
        (pat_bracket, cidrize),
        (pat_thirdoctet_wildcard, third_octet),
        (pat_thirdoctet, third_octet),
        (pat_wildcard, cidrize),
        (pat_two_brackets, two_brackets),
        (pat_fourthoctet, cidrize),
        (pat_squarebrackets, cidrize),
        (pat_squarebrackets_fourth, cidrize),
        (pat_simple, cidrize)
    ]

    for pattern, handler in pattern_handlers:
        matched = re.search(pattern, ip)
        if matched is not None:
            return handler(matched.group(0))

    # doesn't fit into any defined form. Must be a typo somewhere.
    raise CategorisationError

def digest_row(rawvalue):
    """Digest the rawvalue of row."""

    def add_dots(s):
        """We are assuming that the input is of the form X,YYY,ZZZ,AAA
           i.e., that the last three bytes are groups of *three* digits
           separated by commas.  If they weren't, they would probably not
           have been misinterpreted as numbers in the first place by Excel."""
        assert s >= 1000000000
        units = [1000000000, 1000000, 1000, 1]
        my_divide = lambda unit : s / unit % 1000
        return u".".join(map(str, map(my_divide, units)))

    if(type(rawvalue)) == type(1L):
        rawvalue = replace_all(add_dots(rawvalue), reps)
    else:
        rawvalue = replace_all(rawvalue, reps)

    # Remove whitespace and produce an array of (hopefully) readable
    # IP Addresses.
    ss = [x.strip() for x in re.split('[,;&]', rawvalue)]

    cidrizedarray = []

    def valid(i):
        if i is None:
            return False
        if i == u'':
            return False
        return True

    for ip in ss:
        if ip is None or ip is u'':
            continue

        try:
            cidrizedip = categorise_individual_ip(ip)
            for i in cidrizedip:
                if not valid(i):
                    continue
                yield GOOD, i

        except CidrizeError as e:
            yield BAD, ip
        except CategorisationError as e:
            yield BAD, ip
        except:
            print >>sys.stderr, "Unhandled exception in digest_row()!"
            raise

bad_count = 0
def process(sheet, columnnumber):
    def cell_values():
        for row_id in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(column=columnnumber, row=row_id).value
            if cell_value is not None:
                yield (row_id, cell_value)

    def processed_cell_values():
        global bad_count

        # process each row
        for row_id, rawvalue in cell_values():
            cidrized = []
            for status, result in digest_row(rawvalue):
                if status == GOOD:
                    cidrized.append(result)
                else:
                    print >> sys.stderr, result
                    bad_count += 1

            yield (row_id, rawvalue, cidrized)


    dictionary_of_rows_ips = dict([ (row_id, map(str,cidrized)) for row_id, rawvalue,cidrized in processed_cell_values() if cidrized  ])
    return dictionary_of_rows_ips




# run()
def run():
    _, inputf, sheetname, columnnumber = sys.argv
    wb = load_workbook(inputf)
    sheet = wb[sheetname]
    process(sheet, int(columnnumber))

if __name__ == "__main__":
    run()
