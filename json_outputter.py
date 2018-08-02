#!/usr/bin/env python2

# json_outputter.py

# json_outputter, by Chuan Tan <ct538@cam.ac.uk>
#
# Copyright (C) Chuan Tan 2018
#
# This programme is free software; you may redistribute and/or modify
# it under the terms of the Apache Licence, version 2.0.

# Usage: json_outputter.py [inputfile] [sheetname] [institution_col_id]
#                           [country_col_id] [institution_notes_col_id]
#                           [contact_col_id] [ip_col_id]

# Reads an excel file, and processes the IP addresses using standardise-ip.
# It then packages the each row into a JSON object, and ouputs a list of these JSON Objects.
# E.g. JSON Object = {
#       "institution_name" : "University of Cambridge"
#       "country_name"     : "UK"
#       "contacts"     : {"contact_names":"John" , "contact_emails":"john@cam.ac.uk"}
#       "ip_ranges"    : ["123.332.23.2/24","223.225.110.23/32"]
#       }
import sys
from openpyxl import load_workbook
import standardise_ip
import json

SHNAME  = "Sheet1"
INS_COL = 1
COU_COL = 2
INO_COL = 3
CON_COL = 4
IPR_COL = 5

def process_file(inputf, sheetname, institution_col_id, country_col_id, institution_notes_col_id, contact_col_id, ip_col_id):
    wb = load_workbook(inputf)
    sheet = wb[sheetname]
    row_ids_ip_range = standardise_ip.process(sheet,int(ip_col_id))
    JSON_Objects = []
    for row_id in row_ids_ip_range:
        JSON_Object = {
            "institution_name" : sheet.cell(column=int(institution_col_id), row = int(row_id)).value,
            "country_name" : sheet.cell(column=int(country_col_id), row = int(row_id)).value,
            "institution_notes": sheet.cell(column=int(institution_notes_col_id), row = int(row_id)).value,
            "contacts" : [{"contact_name" : sheet.cell(column=int(contact_col_id), row = int(row_id)).value,
                          "contact_email_address" : sheet.cell(column=int(contact_col_id)+1, row = int(row_id)).value}],
            "ip_ranges" : row_ids_ip_range[row_id]
            }
        JSON_Objects.append(JSON_Object)
    print json.dumps(JSON_Objects)

def run():
    if len(sys.argv) == 2:
        _, inputf = sys.argv
        sheetname = SHNAME
        institution_col_id = INS_COL
        country_col_id = COU_COL
        institution_notes_col_id = INO_COL
        contact_col_id = CON_COL
        ip_col_id = IPR_COL
    elif len(sys.argv) == 8:
        _, inputf, sheetname, institution_col_id , country_col_id, institution_notes_col_id, contact_col_id, ip_col_id = sys.argv
    else:
        print >>sys.stderr, "Not enough arguments!"
        sys.exit()
    process_file(inputf, sheetname, institution_col_id , country_col_id, institution_notes_col_id, contact_col_id, ip_col_id)

if __name__ == "__main__":
    run()
